# history:
# 2020/4/12  v1.0  initial
# 2020/4/13  v1.1  output 2 sheets
# 2020/5/11  v1.2  add exception check
# 2020/5/11  v2.0  change output scheme
# 2020/12/17 v2.1  add project sheet
# 2020/12/27 v3.0  re-design
# 2021/2/3   v3.1  chang out scheme
# 2021/6/10  v3.2  work load sheet changed
# 2022/3/24  v3.3  chang out scheme


import os
import sys
import msvcrt
import datetime
import pandas as pd
import numpy as np

import openpyxl
from openpyxl.styles import colors, Font, Color, Border, Side, Alignment, PatternFill

from PyQt5 import QtWidgets
from myform import Ui_Form

import traceback

VERSION = '3.3'
IS_WINDOW = False
IS_WINDOW = True

INX_NAME_WORK_LOAD = 3
INX_NAME_DETAIL_COST = 1
SH_COMPANY = ['上海']
GD_COMPANY = ['广上', '广州', '广深']
NJ_COMPANY = ['南京', '南上']


def print_version(version):
    print('=' * 70)
    print('%s version: %s' % (os.path.basename(__file__), VERSION))
    print('=' * 70)


def wait_any_key():
    print('press any key to exit...')
    msvcrt.getch()


def clean_work_load(df_work_load):
    valid_columns = df_work_load.columns.values[INX_NAME_WORK_LOAD-1:-2]
    df = df_work_load[valid_columns]
    df = df.set_index(valid_columns[2])
    df = df.fillna(0)
    index_name = df.index.name[0]
    df.index.name = index_name
    return df


def clean_detail_cost(df_detail_cost):
    # valid_columns = df_detail_cost.columns.values[INX_NAME_DETAIL_COST:-1]
    valid_columns = df_detail_cost.columns.values[INX_NAME_DETAIL_COST:]
    df = df_detail_cost[valid_columns]
    df = df.set_index(valid_columns[0])
    df = df.fillna(0)
    return df


def check_data(df_work_load, df_detail_cost):
    df = df_work_load.copy()

    df = df_work_load.copy()
    col_attr = df.columns[0]
    s = (df[col_attr] == '研发') | (df[col_attr] == '非研发')
    if s.all() == False:
        txt = '<工时比例>：属性只能为“研发” “非研发”'
        raise Exception(txt)

    col_attr = df.columns[1]
    valid_val = set(SH_COMPANY+GD_COMPANY+NJ_COMPANY)
    all_val = set(df[col_attr].tolist())
    if (valid_val | all_val) != valid_val:
        txt = '<工时比例>：支付归口不为%s' % valid_val
        raise Exception(txt)

    df1 = df.iloc[:, 2:]
    df1 = df1.select_dtypes(exclude=['float64', 'int64'])
    if not df1.empty:
        txt = df1.columns.values
        raise Exception('<工时比例>%s: 有非法值(not float64/int64)' % txt)
        # return '<工时比例>%s: 有非法值(not float64)' % txt

    df1 = df_work_load.sum(axis=1)
    criteria = (df1 < 0.9999) | (df1 > 1.00001)
    if criteria.any():
        txt = df1[criteria].index.values
        raise Exception('<工时比例>%s: 合计不是100%%' % txt)

    col_attr = df.columns[0]
    df1 = df[df[col_attr] != '研发'].iloc[:, -1:]
    criteria = (df1 < 0.9999) | (df1 > 1.00001)
    s_criteria = criteria.T.iloc[0]
    if s_criteria.any():
        txt = df1[s_criteria].index.values
        raise Exception('<工时比例>%s: 非研发类“其他”不为100%%' % txt)

    col_attr = df.columns[0]
    df1 = df[df[col_attr] == '研发'].iloc[:, -1:]
    criteria = (df1 < -0.00001) | (df1 > 0.00001)
    s_criteria = criteria.T.iloc[0]
    if s_criteria.any():
        txt = df1[s_criteria].index.values
        raise Exception('<工时比例>%s: 研发类“其他”不为0%%' % txt)

    df = df_work_load.copy()
    col_attr = df.columns[1]
    loc_txt = '上海'
    for loc in SH_COMPANY:
        df1 = df[df[col_attr] == loc]
        df2 = df1.iloc[:, :-1]
        sr = df2.stack([0, 1, 2])
        sr1 = sr.loc[(slice(None), slice(None), slice(None), ['广东', '南京'])]
        criteria = (sr1 < -0.00001) | (sr1 > 0.00001)
        if criteria.any():
            txt = sr1[criteria].index.values
            raise Exception('<工时比例>%s: 非"%s"项目不为0' % (txt, loc_txt))
    loc_txt = '广东'
    for loc in GD_COMPANY:
        df1 = df[df[col_attr] == loc]
        df2 = df1.iloc[:, :-1]
        sr = df2.stack([0, 1, 2])
        sr1 = sr.loc[(slice(None), slice(None), slice(None), ['上海', '南京'])]
        criteria = (sr1 < -0.00001) | (sr1 > 0.00001)
        if criteria.any():
            txt = sr1[criteria].index.values
            raise Exception('<工时比例>%s: 非"%s"项目不为0' % (txt, loc_txt))
    loc_txt = '南京'
    for loc in NJ_COMPANY:
        df1 = df[df[col_attr] == loc]
        df2 = df1.iloc[:, :-1]
        sr = df2.stack([0, 1, 2])
        sr1 = sr.loc[(slice(None), slice(None), slice(None), ['上海', '广东'])]
        criteria = (sr1 < -0.00001) | (sr1 > 0.00001)
        if criteria.any():
            txt = sr1[criteria].index.values
            raise Exception('<工时比例>%s: 非"%s"项目不为0' % (txt, loc_txt))

    df = df_detail_cost.copy()
    s1 = df.index.names != ['姓名']
    s2 = df.columns.to_list()[0:2] != ['部门', '支付归口']
    if s1 or s2:
        txt = '<费用明细>：列标题应该为“姓名” “部门” “支付归口”'
        raise Exception(txt)

    df1 = df.iloc[:, 2:]
    df1 = df1.select_dtypes(exclude=['float64', 'int64'])
    if not df1.empty:
        txt = df1.columns.values
        raise Exception('<费用明细>%s: 有非法值(not float64/int64)' % txt)

    person_in_cost = set(df_detail_cost.index.values)
    person_in_load = set(df_work_load.index.values)
    person_missed = person_in_load-(person_in_cost & person_in_load)
    if len(person_missed) > 0:
        txt = '<工时比例>人员未出现<费用明细>中: %s' % person_missed
        raise Exception(txt)

    person_missed = person_in_cost-(person_in_cost & person_in_load)
    if len(person_missed) > 0:
        txt = '<费用明细>人员未出现<工时比例>中: %s' % person_missed
        raise Exception(txt)

    if len(person_in_load) != len(df_work_load.index.values):
        txt = '<工时比例>中有重复的员工名'
        raise Exception(txt)

    if len(person_in_cost) != len(df_detail_cost.index.values):
        txt = '<费用明细>中有重复的员工名'
        raise Exception(txt)


# class ExcelDataException(Exception):
#     pass
#     # def __init__(self, msg):
#     #     self.msg = msg

#     # def __str__(self):
#     #     msg = '费用明细表的人员未出现在人员分摊表中: %s' % self.msg
#     #     return(str)


def flatten_work_load_rd(df_work_load_rd, df_detail_cost):
    df = df_work_load_rd.iloc[:, :-1]
    sr = df.stack([0, 1, 2])
    index_v = sr.index.values
    val = sr.to_list()
    new_index = []
    for idx in index_v:
        dept = df_detail_cost.loc[idx[0], '部门']
        loc = df_detail_cost.loc[idx[0], '支付归口']
        new_idx = (idx[0], dept, loc, idx[1], idx[2], idx[3])
        new_index.append(new_idx)

    index = pd.MultiIndex.from_tuples(new_index)
    sr_new = pd.Series(val, index=index)
    sr_new.index.names = ['人员', '部门', '属地', '大类', '项目', '项目归属']
    return sr_new


def nrd_entire(df_work_load_nrd, df_detail_cost):
    nrd_name = df_work_load_nrd.index.values
    df = df_detail_cost.loc[nrd_name]
    df['大类'] = '其他'
    df['项目'] = '0000'
    df['项目归属'] = ''
    df = df.reset_index()
    df = df.set_index(['姓名', '部门', '支付归口', '大类', '项目', '项目归属'])
    df.index.names = ['人员', '部门', '属地', '大类', '项目', '项目归属']
    return df


def rd_entire_before_adj(df_work_load_rd, df_detail_cost):
    sr = flatten_work_load_rd(df_work_load_rd, df_detail_cost)
    df_cost = df_detail_cost.iloc[:, 2:]
    df = df_cost.mul(sr, axis='index', level=0)
    return df


# def rd_entire_after_adj(df_rd_bf_adj):
#     df = df_rd_bf_adj
#     # sum_all = df.sum().sum()
#     sum_sh_emp = df.groupby(['属地']).sum().loc[SH_COMPANY].sum().sum()
#     sum_gd_emp = df.groupby(['属地']).sum().loc[GD_COMPANY].sum().sum()
#     sum_sh_prj = df.groupby(['项目归属']).sum().loc['上海'].sum().sum()
#     sum_gd_prj = df.groupby(['项目归属']).sum().loc['广东'].sum().sum()

#     ratio_sh = sum_sh_emp/sum_sh_prj
#     ratio_gd = sum_gd_emp/sum_gd_prj
#     df_sh_prj = df.loc[(slice(None), slice(None), slice(
#         None), slice(None), slice(None), ['上海']), :].mul(ratio_sh)
#     df_gd_prj = df.loc[(slice(None), slice(None), slice(
#         None), slice(None), slice(None), ['广东']), :].mul(ratio_gd)
#     df_merge = pd.concat([df_sh_prj, df_gd_prj])
#     return df_merge


def format_xlsx(workbook, sheet_name):
    ws = workbook[sheet_name]

    font = Font(name='Arial', size=10, color=colors.BLACK)
    hd_font = Font(name='Arial', size=10, color=colors.BLACK, bold=True)
    thin = Side(border_style="thin")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    alignment = Alignment(horizontal="center", vertical="center")
    fill = PatternFill(fill_type="solid", fgColor="C1CDC1")

    nrows = ws.max_row
    ncols = ws.max_column

    for i in range(nrows):
        for j in range(ncols):
            cell = ws.cell(i+1, j+1)
            cell.number_format = '0.00'
            cell.border = border
            cell.alignment = alignment
            if(cell.font.b):
                cell.font = hd_font
                cell.fill = fill
            else:
                cell.font = font


def write_xlsx(df, xlsx, title, month):
    sheet_name = '%s-%s' % (month, title)
    df.to_excel(xlsx, sheet_name=sheet_name, merge_cells=False)


def show_win_title(myshow, win_title, info):
    txt = '%s [%s]' % (win_title, info)
    myshow.setWindowTitle(txt)


def backend_proc(work_load_file, work_load_sheet, detail_cost_file, detail_cost_sheet, myshow):
    try:
        if(IS_WINDOW):
            win_title = myshow.windowTitle()
            info = '开始'
            show_win_title(myshow, win_title, info)

        xls_file_name = '%s汇总_%s.xlsx' % (
            detail_cost_sheet, datetime.datetime.now().date().strftime('%y%m%d'))

        if(IS_WINDOW):
            info = '读取文件'
            show_win_title(myshow, win_title, info)
        df_work_load = pd.read_excel(
            work_load_file, sheet_name=work_load_sheet,  header=[0, 1, 2])
        df_detail_cost = pd.read_excel(
            detail_cost_file, sheet_name=detail_cost_sheet)
        df_work_load = clean_work_load(df_work_load)
        df_detail_cost = clean_detail_cost(df_detail_cost)
        check_data(df_work_load, df_detail_cost)
        df_detail_cost = df_detail_cost.replace(SH_COMPANY, '上海')
        df_detail_cost = df_detail_cost.replace(GD_COMPANY, '广东')
        df_detail_cost = df_detail_cost.replace(NJ_COMPANY, '南京')

        if(IS_WINDOW):
            info = '数据处理'
            show_win_title(myshow, win_title, info)
        df = df_work_load
        col_attr = df.columns[0]
        col_loc = df.columns[1]
        df_work_load_rd = df[df[col_attr] == '研发'].drop(
            [col_attr, col_loc], axis=1)
        df_work_load_nrd = df[df[col_attr] != '研发'].drop(
            [col_attr, col_loc], axis=1)

        df_nrd = nrd_entire(df_work_load_nrd, df_detail_cost)
        df_rd_bf_adj = rd_entire_before_adj(df_work_load_rd, df_detail_cost)
        # df_rd_af_adj = rd_entire_after_adj(df_rd_bf_adj)

        if(IS_WINDOW):
            info = '生成excel文件'
            show_win_title(myshow, win_title, info)

        xlsx = pd.ExcelWriter(xls_file_name)

        # title = '非研发-全'
        # write_xlsx(df_nrd, xlsx, title, month=work_load_sheet)
        # df = df_nrd.groupby(['部门']).sum()
        # title = '非研发-部门'
        # write_xlsx(df, xlsx, title, month=work_load_sheet)
        # df = df_nrd.groupby(['支付归口']).sum()
        # title = '非研发-属地'
        # write_xlsx(df, xlsx, title, month=work_load_sheet)

        # title = '研发-全-调整前'
        # write_xlsx(df_rd_bf_adj, xlsx, title, month=work_load_sheet)

        # title = '研发-全-调整后'
        # write_xlsx(df_rd_af_adj, xlsx, title, month=work_load_sheet)

        # title = '研发-部门'
        # df = df_rd_af_adj.groupby(['部门']).sum()
        # write_xlsx(df, xlsx, title, month=work_load_sheet)

        # title = '研发-属地'
        # df = df_rd_af_adj.groupby(['属地']).sum()
        # write_xlsx(df, xlsx, title, month=work_load_sheet)

        # title = '研发-大类'
        # df = df_rd_af_adj.groupby(['大类']).sum()
        # write_xlsx(df, xlsx, title, month=work_load_sheet)

        # title = '研发-项目'
        # df = df_rd_af_adj.groupby(['项目', '项目归属']).sum()
        # write_xlsx(df, xlsx, title, month=work_load_sheet)

        # title = '研发-项目归属'
        # df = df_rd_af_adj.groupby(['项目归属']).sum()
        # write_xlsx(df, xlsx, title, month=work_load_sheet)

        # title = '研发-调整前'
        # df = df_rd_bf_adj.groupby(['部门', '属地', '大类', '项目', '项目归属']).sum()
        # write_xlsx(df, xlsx, title, month=work_load_sheet)

        # title = '研发-调整后'
        # df = df_rd_af_adj.groupby(['部门', '属地', '大类', '项目', '项目归属']).sum()
        # write_xlsx(df, xlsx, title, month=work_load_sheet)

        df1 = df_rd_bf_adj
        df2 = df_nrd
        df_whole = pd.concat([df1, df2])
        df = df_whole

        title = '合并'
        df = df.groupby(['部门', '属地', '大类', '项目', '项目归属']).sum()
        df1 = df.loc[(slice(None), ['上海'], slice(
            None), slice(None), ['上海', '']), :]
        df2 = df.loc[(slice(None), ['广东'], slice(
            None), slice(None), ['广东', '']), :]
        df3 = df.loc[(slice(None), ['南京'], slice(
            None), slice(None), ['南京', '']), :]
        df = pd.concat([df1, df2, df3])
        write_xlsx(df, xlsx, title, month=work_load_sheet)

        title = '上海'
        df = df1
        if(len(df)):
            write_xlsx(df, xlsx, title, month=work_load_sheet)

        title = '广东'
        df = df2
        if(len(df)):
            write_xlsx(df, xlsx, title, month=work_load_sheet)

        title = '南京'
        df = df3
        if(len(df)):
            write_xlsx(df, xlsx, title, month=work_load_sheet)

        # df1 = df_rd_af_adj
        # df2 = df_nrd
        # df = pd.concat([df1, df2])

        # title = '合并-调整后'
        # df = df.groupby(['部门', '属地', '大类', '项目', '项目归属']).sum()
        # write_xlsx(df, xlsx, title, month=work_load_sheet)

        df = df_whole
        df = df.loc[~(df == 0).all(axis=1), :]
        # df = df.groupby(['部门', '属地', '大类', '项目', '项目归属']).sum()
        df1 = df.loc[(slice(None), slice(None), ['上海'], slice(
            None), slice(None), ['上海', '']), :]
        df2 = df.loc[(slice(None), slice(None), ['广东'], slice(
            None), slice(None), ['广东', '']), :]
        df3 = df.loc[(slice(None), slice(None), ['南京'], slice(
            None), slice(None), ['南京', '']), :]

        title = '上海-个人'
        df = df1
        if(len(df)):
            write_xlsx(df, xlsx, title, month=work_load_sheet)

        title = '广东-个人'
        df = df2
        if(len(df)):
            write_xlsx(df, xlsx, title, month=work_load_sheet)

        title = '南京-个人'
        df = df3
        if(len(df)):
            write_xlsx(df, xlsx, title, month=work_load_sheet)

        xlsx.close()

        workbook = openpyxl.load_workbook(xls_file_name)
        for sht in workbook.sheetnames:
            if(IS_WINDOW):
                info = '格式-%s' % sht
                show_win_title(myshow, win_title, info)
            format_xlsx(workbook, sheet_name=sht)
        workbook.save(filename=xls_file_name)

    except Exception:
        # return repr(e)
        return traceback.format_exc()

    finally:
        if(IS_WINDOW):
            myshow.setWindowTitle(win_title)

    return(xls_file_name)


class MyWindow(QtWidgets.QWidget, Ui_Form):
    def __init__(self):
        super(MyWindow, self).__init__()
        self.setupUi(self)

    def confirm(self):
        work_load_file = self.txt_work_load.text()
        work_load_sheet = self.txt_work_load_sheet.text()
        detail_cost_file = self.txt_detail_cost.text()
        detail_cost_sheet = self.txt_detail_txt_sheet.text()

        # work_load_file = r'.\工时比例.xlsx'
        # work_load_sheet = '8月'
        # detail_cost_file = r'.\费用明细.xlsx'
        # detail_cost_sheet = '社保'
        # detail_cost_sheet = '公积金'
        # detail_cost_sheet = '工资'

        ret = backend_proc(work_load_file, work_load_sheet,
                           detail_cost_file, detail_cost_sheet, myshow=self)

        if isinstance(ret, str):
            info = '生成文件: %s' % ret
            QtWidgets.QMessageBox.information(self, '信息', info)
        else:
            info = repr(ret)
            QtWidgets.QMessageBox.warning(self, '出错', info)

    def close_win(self):
        self.close()

    def open_work_load(self):
        file_name, _ = QtWidgets.QFileDialog.getOpenFileName(self, '选择文件', '.')
        if file_name:
            self.txt_work_load.setText(file_name)

    def open_detail_cost(self):
        file_name, _ = QtWidgets.QFileDialog.getOpenFileName(self, '选择文件', '.')
        if file_name:
            self.txt_detail_cost.setText(file_name)


def main():
    if IS_WINDOW:
        app = QtWidgets.QApplication(sys.argv)
        myshow = MyWindow()
        myshow.setWindowTitle('cal_cost v%s' % VERSION)
        myshow.show()
        sys.exit(app.exec_())
    else:
        print_version(VERSION)
        work_load_file = r'.\晟矽2021年12月工时分摊表(总表)-20211209 (3)(1)(1).xlsx'
        work_load_sheet = '202102'
        detail_cost_file = r'.\南京2021工时分摊表资料模板(1).xlsx'
        # detail_cost_sheet = '工资'
        detail_cost_sheet = '202101'

        ret = backend_proc(work_load_file, work_load_sheet,
                           detail_cost_file, detail_cost_sheet, myshow=None)

        print(ret)
        # wait_any_key()


if __name__ == '__main__':
    main()
