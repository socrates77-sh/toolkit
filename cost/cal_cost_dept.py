# history:
# 2021/2/4   v1.0  initial
# 2021/6/10  v1.1  work load sheet changed
# 2023/2/10  v1.2  add "江阴","南上"

import os
import sys
import msvcrt
import datetime
import pandas as pd
import numpy as np

import openpyxl
from openpyxl.styles import colors, Font, Color, Border, Side, Alignment, PatternFill

from PyQt5 import QtWidgets
from myform import Ui_Form

import traceback

VERSION = '1.2'
IS_WINDOW = False
IS_WINDOW = True

INX_NAME_WORK_LOAD = 3
INX_NAME_DETAIL_COST = 1
SH_COMPANY = ['上海']
GD_COMPANY = ['广上', '广州', '广深']
NJ_COMPANY = ['南京', '南上']
JY_COMPANY = ['江阴']


def print_version(version):
    print('=' * 70)
    print('%s version: %s' % (os.path.basename(__file__), VERSION))
    print('=' * 70)


def wait_any_key():
    print('press any key to exit...')
    msvcrt.getch()


def clean_work_load(df_work_load):
    valid_columns = df_work_load.columns.values[INX_NAME_WORK_LOAD-1:-2]
    df = df_work_load[valid_columns]
    df = df.set_index(valid_columns[2])
    df = df.fillna(0)
    index_name = df.index.name[0]
    df.index.name = index_name
    return df


def clean_detail_cost(df_detail_cost):
    # valid_columns = df_detail_cost.columns.values[INX_NAME_DETAIL_COST:-1]
    valid_columns = df_detail_cost.columns.values[INX_NAME_DETAIL_COST:]
    df = df_detail_cost[valid_columns]
    df = df.set_index(valid_columns[0])
    df = df.fillna(0)
    return df


def check_data(df_work_load, df_detail_cost):
    df = df_work_load.copy()
    col_attr = df.columns[0]
    s = (df[col_attr] == '研发') | (df[col_attr] == '非研发')
    if s.all() == False:
        txt = '<工时比例>：属性只能为“研发” “非研发”'
        raise Exception(txt)

    col_attr = df.columns[1]
    valid_val = set(SH_COMPANY+GD_COMPANY+NJ_COMPANY+JY_COMPANY)
    all_val = set(df[col_attr].tolist())
    if (valid_val | all_val) != valid_val:
        txt = '<工时比例>：支付归口不为%s' % valid_val
        raise Exception(txt)

    df1 = df.iloc[:, 2:]
    df1 = df1.select_dtypes(exclude=['float64', 'int64'])
    if not df1.empty:
        txt = df1.columns.values
        raise Exception('<工时比例>%s: 有非法值(not float64/int64)' % txt)
        # return '<工时比例>%s: 有非法值(not float64)' % txt

    df1 = df_work_load.sum(axis=1)
    criteria = (df1 < 0.9999) | (df1 > 1.00001)
    if criteria.any():
        txt = df1[criteria].index.values
        raise Exception('<工时比例>%s: 合计不是100%%' % txt)

    col_attr = df.columns[0]
    df1 = df[df[col_attr] != '研发'].iloc[:, -1:]
    criteria = (df1 < 0.9999) | (df1 > 1.00001)
    s_criteria = criteria.T.iloc[0]
    if s_criteria.any():
        txt = df1[s_criteria].index.values
        raise Exception('<工时比例>%s: 非研发类“其他”不为100%%' % txt)

    col_attr = df.columns[0]
    df1 = df[df[col_attr] == '研发'].iloc[:, -1:]
    criteria = (df1 < -0.00001) | (df1 > 0.00001)
    s_criteria = criteria.T.iloc[0]
    if s_criteria.any():
        txt = df1[s_criteria].index.values
        raise Exception('<工时比例>%s: 研发类“其他”不为0%%' % txt)

    df = df_work_load.copy()
    col_attr = df.columns[1]
    loc_txt = '上海'
    for loc in SH_COMPANY:
        df1 = df[df[col_attr] == loc]
        df2 = df1.iloc[:, :-1]
        sr = df2.stack([0, 1, 2])
        sr1 = sr.loc[(slice(None), slice(None),
                      slice(None), ['广东', '南京', '江阴'])]
        criteria = (sr1 < -0.00001) | (sr1 > 0.00001)
        if criteria.any():
            txt = sr1[criteria].index.values
            raise Exception('<工时比例>%s: 非"%s"项目不为0' % (txt, loc_txt))
    loc_txt = '广东'
    for loc in GD_COMPANY:
        df1 = df[df[col_attr] == loc]
        df2 = df1.iloc[:, :-1]
        sr = df2.stack([0, 1, 2])
        sr1 = sr.loc[(slice(None), slice(None),
                      slice(None), ['上海', '南京', '江阴'])]
        criteria = (sr1 < -0.00001) | (sr1 > 0.00001)
        if criteria.any():
            txt = sr1[criteria].index.values
            raise Exception('<工时比例>%s: 非"%s"项目不为0' % (txt, loc_txt))
    loc_txt = '南京'
    for loc in NJ_COMPANY:
        df1 = df[df[col_attr] == loc]
        df2 = df1.iloc[:, :-1]
        sr = df2.stack([0, 1, 2])
        sr1 = sr.loc[(slice(None), slice(None),
                      slice(None), ['上海', '广东', '江阴'])]
        criteria = (sr1 < -0.00001) | (sr1 > 0.00001)
        if criteria.any():
            txt = sr1[criteria].index.values
            raise Exception('<工时比例>%s: 非"%s"项目不为0' % (txt, loc_txt))
    loc_txt = '江阴'
    for loc in JY_COMPANY:
        df1 = df[df[col_attr] == loc]
        df2 = df1.iloc[:, :-1]
        sr = df2.stack([0, 1, 2])
        sr1 = sr.loc[(slice(None), slice(None),
                      slice(None), ['上海', '广东', '南京'])]
        criteria = (sr1 < -0.00001) | (sr1 > 0.00001)
        if criteria.any():
            txt = sr1[criteria].index.values
            raise Exception('<工时比例>%s: 非"%s"项目不为0' % (txt, loc_txt))

    df = df_detail_cost.copy()
    s1 = df.index.names != ['部门']
    s2 = df.columns.to_list()[0:2] != ['属性', '支付归口']
    if s1 or s2:
        txt = '<部门费用>：列标题应该为“部门” “属性” “支付归口”'
        raise Exception(txt)

    df1 = df.iloc[:, 2:]
    df1 = df1.select_dtypes(exclude=['float64', 'int64'])
    if not df1.empty:
        txt = df1.columns.values
        raise Exception('<部门费用>%s: 有非法值(not float64/int64)' % txt)

    col_attr = df.columns[0]
    s = (df[col_attr] == '研发') | (df[col_attr] == '非研发')
    if s.all() == False:
        txt = '<部门费用>：属性只能为“研发” “非研发”'
        raise Exception(txt)

    col_attr = df.columns[1]
    valid_val = set(SH_COMPANY+GD_COMPANY+NJ_COMPANY+JY_COMPANY)
    all_val = set(df[col_attr].tolist())
    if (valid_val | all_val) != valid_val:
        txt = '<部门费用>：支付归口不为%s' % valid_val
        raise Exception(txt)


def adj_ratio_by_proj_loc(df_work_load, loc_txt):
    df = df_work_load.iloc[:, 2:-1]
    sr = df.stack([0, 1, 2])
    sr.index.names = ['人员', '大类', '项目', '项目归属']
    sr = sr.groupby(['大类', '项目', '项目归属']).sum()
    sum_proj_loc = sr.groupby(['项目归属']).sum().sum()
    ratio_by_proj_loc = sr.groupby(['项目归属']).sum()[loc_txt]/sum_proj_loc
    ratio_by_proj_loc = sr.groupby(['项目归属']).sum()[loc_txt]
    # print(sr.loc[(slice(None), slice(None), [loc_txt])].mul(1/ratio_by_proj_loc))
    return sr.loc[(slice(None), slice(None), [loc_txt])].mul(1/ratio_by_proj_loc)


def entire_cost(df_work_load, df_detail_cost):
    df = df_detail_cost.copy()
    df = df.reset_index()
    df = df.set_index(['部门', '属性', '支付归口'])

    index_v = df.index.values

    new_index = []
    new_val = []
    for idx in index_v:
        dept = idx[0]
        loc = idx[2]
        if idx[1] == '非研发':
            val = df.loc[idx].values
            # dept=idx[0]
            # loc = idx[2]
            family = '其他'
            project = '0000'
            proj_loc = ''
            new_idx = (dept, loc, family, project, proj_loc)
            new_index.append(new_idx)
            new_val.append(val)
            # print(val)
        else:
            if loc in SH_COMPANY:
                loc_txt = '上海'
            elif loc in GD_COMPANY:
                loc_txt = '广东'
            else:
                loc_txt = '南京'

            sr = adj_ratio_by_proj_loc(df_work_load, loc_txt)
            for idx_sr in sr.index.values:
                # print(idx_sr)
                ratio = sr[idx_sr]
                family = idx_sr[0]
                project = idx_sr[1]
                proj_loc = idx_sr[2]
                val = df.loc[idx].values*ratio
                # print(idx)
                # print(dept, loc, family, project, proj_loc,val)
                new_idx = (dept, loc, family, project, proj_loc)
                new_index.append(new_idx)
                new_val.append(val)

    index = pd.MultiIndex.from_tuples(new_index)
    df_new = pd.DataFrame(new_val, index=index, columns=df.columns)
    df_new.index.names = ['部门', '属地', '大类', '项目', '项目归属']
    return df_new


def format_xlsx(workbook, sheet_name):
    ws = workbook[sheet_name]

    font = Font(name='Arial', size=10, color=colors.BLACK)
    hd_font = Font(name='Arial', size=10, color=colors.BLACK, bold=True)
    thin = Side(border_style="thin")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    alignment = Alignment(horizontal="center", vertical="center")
    fill = PatternFill(fill_type="solid", fgColor="C1CDC1")

    nrows = ws.max_row
    ncols = ws.max_column

    for i in range(nrows):
        for j in range(ncols):
            cell = ws.cell(i+1, j+1)
            cell.number_format = '0.00'
            cell.border = border
            cell.alignment = alignment
            if(cell.font.b):
                cell.font = hd_font
                cell.fill = fill
            else:
                cell.font = font


def write_xlsx(df, xlsx, title, month):
    sheet_name = '%s-%s' % (month, title)
    df.to_excel(xlsx, sheet_name=sheet_name, merge_cells=False)


def show_win_title(myshow, win_title, info):
    txt = '%s [%s]' % (win_title, info)
    myshow.setWindowTitle(txt)


def backend_proc(work_load_file, work_load_sheet, detail_cost_file, detail_cost_sheet, myshow):
    try:
        if(IS_WINDOW):
            win_title = myshow.windowTitle()
            info = '开始'
            show_win_title(myshow, win_title, info)

        xls_file_name = '%s汇总_%s.xlsx' % (
            detail_cost_sheet, datetime.datetime.now().date().strftime('%y%m%d'))

        if(IS_WINDOW):
            info = '读取文件'
            show_win_title(myshow, win_title, info)

        df_work_load = pd.read_excel(
            work_load_file, sheet_name=work_load_sheet,  header=[0, 1, 2])
        df_detail_cost = pd.read_excel(
            detail_cost_file, sheet_name=detail_cost_sheet)
        df_work_load = clean_work_load(df_work_load)
        df_detail_cost = clean_detail_cost(df_detail_cost)
        check_data(df_work_load, df_detail_cost)

        if(IS_WINDOW):
            info = '数据处理'
            show_win_title(myshow, win_title, info)

        if(IS_WINDOW):
            info = '生成excel文件'
            show_win_title(myshow, win_title, info)

        xlsx = pd.ExcelWriter(xls_file_name)

        title = '合并'
        df = entire_cost(df_work_load, df_detail_cost)
        write_xlsx(df, xlsx, title, month=work_load_sheet)

        title = '上海'
        df1 = df.loc[(slice(None), SH_COMPANY), :]
        if not df1.empty:
            write_xlsx(df1, xlsx, title, month=work_load_sheet)

        title = '广东'
        df1 = df.loc[(slice(None), GD_COMPANY), :]
        if not df1.empty:
            write_xlsx(df1, xlsx, title, month=work_load_sheet)

        title = '南京'
        df1 = df.loc[(slice(None), NJ_COMPANY), :]
        if not df1.empty:
            write_xlsx(df1, xlsx, title, month=work_load_sheet)

        title = '江阴'
        df1 = df.loc[(slice(None), JY_COMPANY), :]
        if not df1.empty:
            write_xlsx(df1, xlsx, title, month=work_load_sheet)

        xlsx.close()

        workbook = openpyxl.load_workbook(xls_file_name)
        for sht in workbook.sheetnames:
            if(IS_WINDOW):
                info = '格式-%s' % sht
                show_win_title(myshow, win_title, info)
            format_xlsx(workbook, sheet_name=sht)
        workbook.save(filename=xls_file_name)

    except Exception:
        # return repr(e)
        return traceback.format_exc()

    finally:
        if(IS_WINDOW):
            myshow.setWindowTitle(win_title)

    return(xls_file_name)


class MyWindow(QtWidgets.QWidget, Ui_Form):
    def __init__(self):
        super(MyWindow, self).__init__()
        self.setupUi(self)

    def confirm(self):
        work_load_file = self.txt_work_load.text()
        work_load_sheet = self.txt_work_load_sheet.text()
        detail_cost_file = self.txt_detail_cost.text()
        detail_cost_sheet = self.txt_detail_txt_sheet.text()

        # work_load_file = r'.\工时比例.xlsx'
        # work_load_sheet = '8月'
        # detail_cost_file = r'.\费用明细.xlsx'
        # detail_cost_sheet = '社保'
        # detail_cost_sheet = '公积金'
        # detail_cost_sheet = '工资'

        ret = backend_proc(work_load_file, work_load_sheet,
                           detail_cost_file, detail_cost_sheet, myshow=self)

        if isinstance(ret, str):
            info = '生成文件: %s' % ret
            QtWidgets.QMessageBox.information(self, '信息', info)
        else:
            info = repr(ret)
            QtWidgets.QMessageBox.warning(self, '出错', info)

    def close_win(self):
        self.close()

    def open_work_load(self):
        file_name, _ = QtWidgets.QFileDialog.getOpenFileName(self, '选择文件', '.')
        if file_name:
            self.txt_work_load.setText(file_name)

    def open_detail_cost(self):
        file_name, _ = QtWidgets.QFileDialog.getOpenFileName(self, '选择文件', '.')
        if file_name:
            self.txt_detail_cost.setText(file_name)


def main():
    if IS_WINDOW:
        app = QtWidgets.QApplication(sys.argv)
        myshow = MyWindow()
        myshow.setWindowTitle('cal_cost_dept v%s' % VERSION)
        myshow.show()
        sys.exit(app.exec_())
    else:
        print_version(VERSION)

        work_load_file = r'.\工时分摊表202110调整.xlsx'
        work_load_sheet = '202110'
        # detail_cost_file = r'.\部门费用2.xlsx'
        # detail_cost_sheet = '日常费用'
        detail_cost_file = r'.\部门费用1.xlsx'
        detail_cost_sheet = '10月份'

        ret = backend_proc(work_load_file, work_load_sheet,
                           detail_cost_file, detail_cost_sheet, myshow=None)

        print(ret)
        # wait_any_key()


if __name__ == '__main__':
    main()
